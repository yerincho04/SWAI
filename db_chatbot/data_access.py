#!/usr/bin/env python3
"""Data access layer for normalized brand tables."""

from __future__ import annotations

import json
import os
import re
import sqlite3
import zipfile
from difflib import SequenceMatcher
from collections import defaultdict
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET


def _normalize_brand_key(text: str) -> str:
    lowered = str(text).strip().lower()
    no_space = "".join(lowered.split())
    # Keep alnum + Hangul only for stable matching.
    return re.sub(r"[^0-9a-z가-힣]", "", no_space)


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _format_int(value: int | None) -> str:
    if value is None:
        return "N/A"
    return f"{value:,}"


def _format_krw(value: int | None) -> str:
    if value is None:
        return "not disclosed in source data"
    return f"{value:,} KRW"


def _format_pct(decimal_ratio: float | None) -> str:
    if decimal_ratio is None:
        return "N/A"
    return f"{decimal_ratio * 100:.2f}%"


def _load_json(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_json_if_exists(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        data = _load_json(path)
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [r for r in data if isinstance(r, dict)]


def _safe_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def _parse_range_mid(value: Any) -> int | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    if not nums:
        return None
    try:
        values = [float(n) for n in nums]
    except Exception:
        return None
    return int(round(sum(values) / len(values)))


def _load_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return _load_xlsx_rows_via_xml(path)

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if not wb.sheetnames:
            return []
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(v).strip() if v is not None else "" for v in header_row]
        out: list[dict[str, Any]] = []
        for row in rows_iter:
            rec = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
            if any(v is not None and str(v).strip() != "" for v in rec.values()):
                out.append(rec)
        return out
    except Exception:
        return _load_xlsx_rows_via_xml(path)


def _xlsx_col_to_index(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha()).upper()
    value = 0
    for ch in letters:
        value = (value * 26) + (ord(ch) - ord("A") + 1)
    return max(0, value - 1)


def _coerce_xlsx_value(text: str | None) -> Any:
    if text is None:
        return None
    raw = str(text).strip()
    if raw == "":
        return None
    if re.fullmatch(r"-?\d+", raw):
        try:
            return int(raw)
        except Exception:
            return raw
    if re.fullmatch(r"-?\d+\.\d+", raw):
        try:
            num = float(raw)
            if num.is_integer():
                return int(num)
            return num
        except Exception:
            return raw
    return raw


def _load_xlsx_rows_via_xml(path: Path) -> list[dict[str, Any]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
        "docrel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    def _resolve_first_sheet(zf: zipfile.ZipFile) -> str | None:
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = workbook_xml.find("main:sheets", ns)
        if sheets is None:
            return None
        first_sheet = sheets.find("main:sheet", ns)
        if first_sheet is None:
            return None
        rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if not rel_id:
            return "xl/worksheets/sheet1.xml"
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        for rel in rels_xml.findall("rel:Relationship", ns):
            if rel.attrib.get("Id") == rel_id:
                target = rel.attrib.get("Target")
                if target:
                    target = target.lstrip("/")
                    if not target.startswith("xl/"):
                        target = f"xl/{target}"
                    return target
        return "xl/worksheets/sheet1.xml"

    def _load_shared_strings(zf: zipfile.ZipFile) -> list[str]:
        if "xl/sharedStrings.xml" not in zf.namelist():
            return []
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        values: list[str] = []
        for si in root.findall("main:si", ns):
            text_parts = [node.text or "" for node in si.findall(".//main:t", ns)]
            values.append("".join(text_parts))
        return values

    try:
        with zipfile.ZipFile(path) as zf:
            sheet_path = _resolve_first_sheet(zf)
            if not sheet_path or sheet_path not in zf.namelist():
                return []
            shared_strings = _load_shared_strings(zf)
            root = ET.fromstring(zf.read(sheet_path))
    except Exception:
        return []

    rows_by_index: list[list[Any]] = []
    sheet_data = root.find("main:sheetData", ns)
    if sheet_data is None:
        return []

    for row in sheet_data.findall("main:row", ns):
        values: dict[int, Any] = {}
        max_index = -1
        for cell in row.findall("main:c", ns):
            ref = cell.attrib.get("r", "")
            col_idx = _xlsx_col_to_index(ref)
            max_index = max(max_index, col_idx)
            cell_type = cell.attrib.get("t")
            if cell_type == "inlineStr":
                value = "".join(node.text or "" for node in cell.findall(".//main:t", ns))
            else:
                value_node = cell.find("main:v", ns)
                raw_value = value_node.text if value_node is not None else None
                if cell_type == "s" and raw_value is not None:
                    try:
                        value = shared_strings[int(raw_value)]
                    except Exception:
                        value = raw_value
                else:
                    value = _coerce_xlsx_value(raw_value)
            values[col_idx] = value
        if max_index >= 0:
            rows_by_index.append([values.get(i) for i in range(max_index + 1)])

    if not rows_by_index:
        return []
    headers = [str(v).strip() if v is not None else "" for v in rows_by_index[0]]
    out: list[dict[str, Any]] = []
    for row in rows_by_index[1:]:
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            out.append(rec)
    return out


class SelectedTableSource:
    """Abstract loader for the five raw selected source tables."""

    def load_brand_list_info(self) -> list[dict[str, Any]]:
        raise NotImplementedError

    def load_brand_frcs_stats(self) -> list[dict[str, Any]]:
        raise NotImplementedError

    def load_brand_fntn_stats(self) -> list[dict[str, Any]]:
        raise NotImplementedError

    def load_brand_brand_stats(self) -> list[dict[str, Any]]:
        raise NotImplementedError

    def load_brand_interior_cost(self) -> list[dict[str, Any]]:
        raise NotImplementedError


class ExcelSelectedTableSource(SelectedTableSource):
    """Active table-style source backed by one Excel file per selected dataset."""

    def __init__(self, root: Path | str) -> None:
        self.root = Path(root)

    def _load(self, relative_path: str) -> list[dict[str, Any]]:
        return _load_xlsx_rows(self.root / relative_path)

    def load_brand_list_info(self) -> list[dict[str, Any]]:
        return self._load("brand_list_info/output/brand_list_info.xlsx")

    def load_brand_frcs_stats(self) -> list[dict[str, Any]]:
        return self._load("brand_frcs_stats/output/brand_frcs_stats.xlsx")

    def load_brand_fntn_stats(self) -> list[dict[str, Any]]:
        return self._load("brand_fntn_stats/output/brand_fntn_stats.xlsx")

    def load_brand_brand_stats(self) -> list[dict[str, Any]]:
        return self._load("brand_brand_stats/output/brand_brand_stats.xlsx")

    def load_brand_interior_cost(self) -> list[dict[str, Any]]:
        return self._load("brand_interior_cost/output/brand_interior_cost.xlsx")


class JsonSelectedTableSource(SelectedTableSource):
    """Backward-compatible fallback source backed by selected JSON files."""

    def __init__(self, root: Path | str) -> None:
        self.root = Path(root)

    def _load(self, relative_path: str) -> list[dict[str, Any]]:
        return _load_json_if_exists(self.root / relative_path)

    def load_brand_list_info(self) -> list[dict[str, Any]]:
        return self._load("brand_list_info/output/brand_list_info_selected.json")

    def load_brand_frcs_stats(self) -> list[dict[str, Any]]:
        return self._load("brand_frcs_stats/output/brand_frcs_stats_selected.json")

    def load_brand_fntn_stats(self) -> list[dict[str, Any]]:
        return self._load("brand_fntn_stats/output/brand_fntn_stats_selected.json")

    def load_brand_brand_stats(self) -> list[dict[str, Any]]:
        return self._load("brand_brand_stats/output/brand_brand_stats_selected.json")

    def load_brand_interior_cost(self) -> list[dict[str, Any]]:
        return self._load("brand_interior_cost/output/brand_interior_cost_selected.json")


class InMemorySelectedTableSource(SelectedTableSource):
    """Table source backed by already-loaded row lists.

    This is the intended shape for a future database-backed adapter:
    execute SELECT queries externally, then pass the resulting rows in here.
    """

    def __init__(
        self,
        brand_list_info_rows: list[dict[str, Any]] | None = None,
        brand_frcs_stats_rows: list[dict[str, Any]] | None = None,
        brand_fntn_stats_rows: list[dict[str, Any]] | None = None,
        brand_brand_stats_rows: list[dict[str, Any]] | None = None,
        brand_interior_cost_rows: list[dict[str, Any]] | None = None,
    ) -> None:
        self.brand_list_info_rows = list(brand_list_info_rows or [])
        self.brand_frcs_stats_rows = list(brand_frcs_stats_rows or [])
        self.brand_fntn_stats_rows = list(brand_fntn_stats_rows or [])
        self.brand_brand_stats_rows = list(brand_brand_stats_rows or [])
        self.brand_interior_cost_rows = list(brand_interior_cost_rows or [])

    def load_brand_list_info(self) -> list[dict[str, Any]]:
        return list(self.brand_list_info_rows)

    def load_brand_frcs_stats(self) -> list[dict[str, Any]]:
        return list(self.brand_frcs_stats_rows)

    def load_brand_fntn_stats(self) -> list[dict[str, Any]]:
        return list(self.brand_fntn_stats_rows)

    def load_brand_brand_stats(self) -> list[dict[str, Any]]:
        return list(self.brand_brand_stats_rows)

    def load_brand_interior_cost(self) -> list[dict[str, Any]]:
        return list(self.brand_interior_cost_rows)


class DatabaseSelectedTableSource(SelectedTableSource):
    """Database-backed table source for future SQL integration.

    This adapter currently targets SQLite via the stdlib `sqlite3` module.
    It intentionally keeps the contract simple: each method returns the full
    table as a list of dict rows, matching the Excel/JSON source shapes.
    """

    DEFAULT_TABLE_MAP = {
        "brand_list_info": "brand_list_info",
        "brand_frcs_stats": "brand_frcs_stats",
        "brand_fntn_stats": "brand_fntn_stats",
        "brand_brand_stats": "brand_brand_stats",
        "brand_interior_cost": "brand_interior_cost",
    }

    def __init__(
        self,
        db_path: Path | str,
        table_map: dict[str, str] | None = None,
    ) -> None:
        self.db_path = Path(db_path)
        self.table_map = {**self.DEFAULT_TABLE_MAP, **(table_map or {})}

    def _fetch_all(self, logical_table_name: str) -> list[dict[str, Any]]:
        table_name = self.table_map[logical_table_name]
        query = f"SELECT * FROM {table_name} WHERE 1=1"
        try:
            conn = sqlite3.connect(str(self.db_path))
            conn.row_factory = sqlite3.Row
        except Exception as exc:
            raise RuntimeError(f"Failed to open database at '{self.db_path}': {exc}") from exc

        try:
            rows = conn.execute(query).fetchall()
        except Exception as exc:
            raise RuntimeError(
                f"Failed to read logical table '{logical_table_name}' from database table '{table_name}': {exc}"
            ) from exc
        finally:
            conn.close()
        return [dict(row) for row in rows]

    def load_brand_list_info(self) -> list[dict[str, Any]]:
        return self._fetch_all("brand_list_info")

    def load_brand_frcs_stats(self) -> list[dict[str, Any]]:
        return self._fetch_all("brand_frcs_stats")

    def load_brand_fntn_stats(self) -> list[dict[str, Any]]:
        return self._fetch_all("brand_fntn_stats")

    def load_brand_brand_stats(self) -> list[dict[str, Any]]:
        return self._fetch_all("brand_brand_stats")

    def load_brand_interior_cost(self) -> list[dict[str, Any]]:
        return self._fetch_all("brand_interior_cost")


def create_selected_table_source(
    source_mode: str,
    api_data_root: Path | str,
    db_path: Path | str | None = None,
) -> SelectedTableSource:
    root = Path(api_data_root)
    if source_mode == "api_selected":
        return JsonSelectedTableSource(root)
    if source_mode == "db_selected":
        resolved_db_path = Path(
            db_path or os.getenv("DB_SELECTED_PATH") or "db_chatbot/selected_tables.sqlite3"
        )
        return DatabaseSelectedTableSource(resolved_db_path)
    return ExcelSelectedTableSource(root)


class BrandResolutionError(ValueError):
    def __init__(
        self,
        query_text: str,
        status: str,
        candidates: list[dict[str, Any]] | None = None,
        reason: str | None = None,
    ) -> None:
        self.query_text = query_text
        self.status = status
        self.candidates = candidates or []
        self.reason = reason or ""
        if status == "ambiguous":
            cand_text = ", ".join(
                f"{c['brand_name']}({c['confidence']:.2f})" for c in self.candidates
            )
            msg = f"브랜드명 '{query_text}'이(가) 모호합니다. 다음 후보 중 선택해 주세요: {cand_text}"
        else:
            msg = f"브랜드 '{query_text}'을(를) 찾을 수 없습니다. 입력을 다시 확인해 주세요."
        super().__init__(msg)

    def to_payload(self) -> dict[str, Any]:
        return {
            "error_type": "brand_resolution",
            "resolution_status": self.status,
            "query_text": self.query_text,
            "candidates": self.candidates,
            "reason": self.reason,
            "error": str(self),
        }


class BrandDataStore:
    def __init__(
        self,
        build_dir: Path | str = Path("db_chatbot/build"),
        source_mode: str = "excel_selected",
        api_data_root: Path | str = Path("db_chatbot/api_data"),
        db_path: Path | str | None = None,
        selected_source: SelectedTableSource | None = None,
    ) -> None:
        self.build_dir = Path(build_dir)
        self.api_data_root = Path(api_data_root)
        self.source_mode = source_mode
        self.db_path = Path(db_path) if db_path is not None else None
        self.selected_source = selected_source or create_selected_table_source(
            source_mode=source_mode,
            api_data_root=self.api_data_root,
            db_path=self.db_path,
        )

        self._api_mnno_to_id: dict[str, int] = {}
        self._api_name_to_id: dict[str, int] = {}
        if self.source_mode in {"api_selected", "excel_selected", "db_selected"} or selected_source is not None:
            self.brand_master, self.brand_year_stats, self.brand_store_types, self.brand_store_type_costs = (
                self._load_core_from_selected_source()
            )
        else:
            self.brand_master = _load_json(self.build_dir / "brand_master.json")
            self.brand_year_stats = _load_json(self.build_dir / "brand_year_stats.json")
            self.brand_store_types = _load_json(self.build_dir / "brand_store_types.json")
            self.brand_store_type_costs = _load_json(self.build_dir / "brand_store_type_costs.json")

        self.master_by_id = {row["brand_id"]: row for row in self.brand_master}
        self.brand_ids_by_lower_name: dict[str, list[int]] = defaultdict(list)
        self.brand_ids_by_normalized_name: dict[str, list[int]] = defaultdict(list)
        self.brand_search_signals: dict[int, list[str]] = defaultdict(list)
        for row in self.brand_master:
            brand_id = row["brand_id"]
            brand_name = row["brand_name"]
            company_name = row.get("company_name") or ""

            self.brand_ids_by_lower_name[brand_name.lower()].append(brand_id)
            self.brand_ids_by_normalized_name[_normalize_brand_key(brand_name)].append(brand_id)

            # Search signals are generated from data only (no hardcoded alias map).
            signals = [brand_name]
            if company_name:
                signals.append(company_name)
            normalized_signals = []
            for signal in signals:
                norm = _normalize_brand_key(signal)
                if norm:
                    normalized_signals.append(norm)
            self.brand_search_signals[brand_id] = sorted(set(normalized_signals))

        self.year_stats_by_brand: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in self.brand_year_stats:
            self.year_stats_by_brand[row["brand_id"]].append(row)
        for brand_id, rows in self.year_stats_by_brand.items():
            self.year_stats_by_brand[brand_id] = sorted(rows, key=lambda r: r["year"])

        self.store_types_by_brand: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in self.brand_store_types:
            self.store_types_by_brand[row["brand_id"]].append(row)

        self.costs_by_brand_year_type: dict[tuple[int, int, str], dict[str, int]] = defaultdict(dict)
        for row in self.brand_store_type_costs:
            key = (row["brand_id"], row["year"], row["store_type"])
            self.costs_by_brand_year_type[key][row["cost_category"]] = row["cost_amount_krw"]

        # Resolver tuning knobs (override by env vars when needed).
        self.resolver_top_k = int(os.getenv("RESOLVER_TOP_K", "3"))
        self.resolver_high_confidence = float(os.getenv("RESOLVER_HIGH_CONF", "0.90"))
        self.resolver_high_margin = float(os.getenv("RESOLVER_HIGH_MARGIN", "0.08"))
        self.resolver_ambiguous_min = float(os.getenv("RESOLVER_AMBIG_MIN", "0.75"))
        self.resolver_llm_min = float(os.getenv("RESOLVER_LLM_MIN", "0.70"))

        self.brand_list_info_latest_by_brand: dict[int, dict[str, Any]] = {}
        self.frcs_extra_by_brand_year: dict[int, dict[int, dict[str, Any]]] = defaultdict(dict)
        self.fntn_extra_by_brand_year: dict[int, dict[int, dict[str, Any]]] = defaultdict(dict)
        self.brand_extra_by_brand_year: dict[int, dict[int, dict[str, Any]]] = defaultdict(dict)
        self.interior_extra_by_brand_year: dict[int, dict[int, dict[str, Any]]] = defaultdict(dict)
        self._load_api_data_enrichment()

    def _resolve_api_brand_id(self, row: dict[str, Any]) -> int | None:
        brand_id = _safe_int(row.get("brand_id"))
        if brand_id is not None:
            return brand_id
        mnno = str(row.get("brandMnno") or "").strip()
        if mnno and mnno in self._api_mnno_to_id:
            return self._api_mnno_to_id[mnno]
        name = str(row.get("brandNm") or "").strip()
        if name and name in self._api_name_to_id:
            return self._api_name_to_id[name]
        return None

    def _load_core_from_selected_source(
        self,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
        brand_list_rows = self.selected_source.load_brand_list_info()
        frcs_rows = self.selected_source.load_brand_frcs_stats()
        fntn_rows = self.selected_source.load_brand_fntn_stats()
        interior_rows = self.selected_source.load_brand_interior_cost()

        # Stable id map from brand_list_info selected rows.
        next_id = 1
        for row in brand_list_rows:
            mnno = str(row.get("brandMnno") or "").strip()
            name = str(row.get("brandNm") or "").strip()
            existing = None
            if mnno and mnno in self._api_mnno_to_id:
                existing = self._api_mnno_to_id[mnno]
            elif name and name in self._api_name_to_id:
                existing = self._api_name_to_id[name]
            if existing is None:
                existing = next_id
                next_id += 1
            if mnno and mnno not in self._api_mnno_to_id:
                self._api_mnno_to_id[mnno] = existing
            if name and name not in self._api_name_to_id:
                self._api_name_to_id[name] = existing

        master_by_id: dict[int, dict[str, Any]] = {}
        for row in brand_list_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None:
                continue
            rec = {
                "brand_id": brand_id,
                "brand_name": row.get("brandNm"),
                "company_name": row.get("corpNm"),
                "category_main": row.get("indutyLclasNm"),
                "category_sub": row.get("indutyMlsfcNm"),
                "franchise_start_date": row.get("jngBizStrtDate"),
            }
            master_by_id[brand_id] = self._pick_richer_row(master_by_id.get(brand_id), rec)
        brand_master = sorted(master_by_id.values(), key=lambda r: r["brand_id"])

        # brand_year_stats from frcs stats.
        rows_by_brand_year: dict[tuple[int, int], dict[str, Any]] = {}
        grouped_years: dict[int, list[dict[str, Any]]] = defaultdict(list)
        for row in frcs_rows:
            brand_id = self._resolve_api_brand_id(row)
            year = _safe_int(row.get("yr"))
            if brand_id is None or year is None:
                continue
            store_count = _safe_int(row.get("frcsCnt")) or 0
            new_stores = _safe_int(row.get("newFrcsRgsCnt")) or 0
            closed_stores = _safe_int(row.get("ctrtCncltnCnt")) or 0
            rec = {
                "brand_id": brand_id,
                "year": year,
                "store_count": store_count,
                "new_stores": new_stores,
                "closed_stores": closed_stores,
                "avg_sales_krw": _safe_int(row.get("avrgSlsAmt")),
                "net_store_change": new_stores - closed_stores,
                "store_growth_rate": 0.0,
                "closure_rate": (closed_stores / store_count) if store_count > 0 else 0.0,
                "churn_rate": ((new_stores + closed_stores) / store_count) if store_count > 0 else 0.0,
            }
            key = (brand_id, year)
            rows_by_brand_year[key] = self._pick_richer_row(rows_by_brand_year.get(key), rec)

        for rec in rows_by_brand_year.values():
            grouped_years[rec["brand_id"]].append(rec)
        brand_year_stats: list[dict[str, Any]] = []
        for rows in grouped_years.values():
            rows.sort(key=lambda r: r["year"])
            for idx, row in enumerate(rows):
                net = row["net_store_change"] or 0
                if idx > 0:
                    prev_store_count = float(rows[idx - 1]["store_count"] or 0)
                else:
                    prev_store_count = float((row["store_count"] or 0) - net)
                row["store_growth_rate"] = (net / prev_store_count) if prev_store_count > 0 else 0.0
                brand_year_stats.append(row)
        brand_year_stats.sort(key=lambda r: (r["brand_id"], r["year"]))

        # brand_store_types from interior.
        type_by_brand: dict[int, dict[str, Any]] = {}
        for row in interior_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None:
                continue
            rec = {
                "brand_id": brand_id,
                "store_type": "Standard",
                "standard_area_pyeong": (
                    float(row["storCrtraAr"]) if row.get("storCrtraAr") is not None else None
                ),
            }
            type_by_brand[brand_id] = self._pick_richer_row(type_by_brand.get(brand_id), rec)
        brand_store_types = sorted(type_by_brand.values(), key=lambda r: (r["brand_id"], r["store_type"]))

        # brand_store_type_costs from fntn + interior midpoints.
        interior_mid_by_brand_year: dict[tuple[int, int], int] = {}
        for row in interior_rows:
            brand_id = self._resolve_api_brand_id(row)
            year = _safe_int(row.get("jngBizCrtraYr"))
            if brand_id is None or year is None:
                continue
            mid = _parse_range_mid(row.get("intrrAmtScopeVal"))
            if mid is not None:
                interior_mid_by_brand_year[(brand_id, year)] = mid

        cost_rows: list[dict[str, Any]] = []
        for row in fntn_rows:
            brand_id = self._resolve_api_brand_id(row)
            year = _safe_int(row.get("yr"))
            if brand_id is None or year is None:
                continue
            mappings = [
                ("initial_fee", _safe_int(row.get("jngBzmnJngAmt"))),
                ("education", _safe_int(row.get("jngBzmnEduAmt"))),
                ("other", _safe_int(row.get("jngBzmnEtcAmt"))),
                ("guarantee", _safe_int(row.get("jngBzmnAssrncAmt"))),
                ("total_initial_cost", _safe_int(row.get("smtnAmt"))),
            ]
            interior_mid = interior_mid_by_brand_year.get((brand_id, year))
            if interior_mid is not None:
                mappings.append(("interior", interior_mid))
            for cat, amount in mappings:
                if amount is None:
                    continue
                cost_rows.append(
                    {
                        "brand_id": brand_id,
                        "year": year,
                        "store_type": "Standard",
                        "cost_category": cat,
                        "cost_amount_krw": amount,
                    }
                )
        dedup_cost: dict[tuple[int, int, str, str], dict[str, Any]] = {}
        for row in cost_rows:
            key = (row["brand_id"], row["year"], row["store_type"], row["cost_category"])
            dedup_cost[key] = row
        brand_store_type_costs = sorted(
            dedup_cost.values(),
            key=lambda r: (r["brand_id"], r["year"], r["store_type"], r["cost_category"]),
        )
        return brand_master, brand_year_stats, brand_store_types, brand_store_type_costs

    def _llm_resolve_brand(self, brand_query: str) -> dict[str, Any] | None:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return None
        try:
            from openai import OpenAI
        except Exception:
            return None

        model = os.getenv("RESOLVER_MODEL", "gpt-4.1-mini")
        candidates = [row["brand_name"] for row in self.brand_master]

        prompt = (
            "You are a brand name resolver.\n"
            "Given a user mention and candidate brand names, choose the single best match.\n"
            "Return strict JSON only with keys: brand_name, confidence, reason.\n"
            "If no reliable match, set brand_name to null.\n"
            f"user_mention: {brand_query}\n"
            f"candidates: {json.dumps(candidates, ensure_ascii=False)}"
        )
        try:
            client = OpenAI(api_key=api_key)
            resp = client.chat.completions.create(
                model=model,
                temperature=0,
                response_format={"type": "json_object"},
                messages=[{"role": "user", "content": prompt}],
            )
            text = (resp.choices[0].message.content or "").strip()
            parsed = json.loads(text)
            brand_name = parsed.get("brand_name")
            confidence = float(parsed.get("confidence", 0.0))
            reason = str(parsed.get("reason", "llm_fallback"))
            if not brand_name:
                return None
            target = next((r for r in self.brand_master if r["brand_name"] == brand_name), None)
            if not target:
                return None
            return {
                "brand_id": target["brand_id"],
                "brand_name": target["brand_name"],
                "confidence": max(0.0, min(confidence, 1.0)),
                "reason": reason,
            }
        except Exception:
            return None

    def _resolve_brand_candidate(self, brand_query: str, top_k: int = 3) -> dict[str, Any]:
        q = brand_query.strip()
        if not q:
            raise ValueError("브랜드명이 비어 있습니다.")

        q_norm = _normalize_brand_key(q)
        if not q_norm:
            return {
                "status": "not_found",
                "query_text": brand_query,
                "match": None,
                "candidates": [],
                "reason": "empty_after_normalization",
            }

        exact_ids = self.brand_ids_by_lower_name.get(q.lower(), [])
        if len(exact_ids) == 1:
            bid = exact_ids[0]
            return {
                "status": "resolved",
                "query_text": brand_query,
                "match": {"brand_id": bid, "brand_name": self.master_by_id[bid]["brand_name"], "confidence": 1.0},
                "candidates": [
                    {
                        "brand_id": bid,
                        "brand_name": self.master_by_id[bid]["brand_name"],
                        "confidence": 1.0,
                        "stage": "exact",
                    }
                ],
                "reason": "exact_match",
            }

        normalized_ids = self.brand_ids_by_normalized_name.get(q_norm, [])
        if len(normalized_ids) == 1:
            bid = normalized_ids[0]
            return {
                "status": "resolved",
                "query_text": brand_query,
                "match": {"brand_id": bid, "brand_name": self.master_by_id[bid]["brand_name"], "confidence": 0.98},
                "candidates": [
                    {
                        "brand_id": bid,
                        "brand_name": self.master_by_id[bid]["brand_name"],
                        "confidence": 0.98,
                        "stage": "normalized_exact",
                    }
                ],
                "reason": "normalized_exact_match",
            }

        scored = []
        for brand_id, signals in self.brand_search_signals.items():
            best_score = 0.0
            for signal in signals:
                if q_norm == signal:
                    score = 0.98
                elif q_norm in signal or signal in q_norm:
                    # Strong substring match from data-driven signals.
                    score = 0.93
                else:
                    score = _similarity(q_norm, signal)
                if score > best_score:
                    best_score = score
            scored.append((brand_id, best_score))

        scored.sort(key=lambda x: x[1], reverse=True)
        effective_top_k = max(top_k or self.resolver_top_k, 2)
        top = scored[:effective_top_k]
        if not top:
            return {
                "status": "not_found",
                "query_text": brand_query,
                "match": None,
                "candidates": [],
                "reason": "no_candidates",
            }

        top1_id, top1_score = top[0]
        top2_score = top[1][1] if len(top) > 1 else 0.0
        raw_candidates = [
            {
                "brand_id": bid,
                "brand_name": self.master_by_id[bid]["brand_name"],
                "confidence": round(score, 4),
                "stage": "fuzzy_rank",
            }
            for bid, score in top[:effective_top_k]
        ]
        # Hide non-informative candidates when all scores are effectively zero.
        candidates = [c for c in raw_candidates if c["confidence"] > 0.01]

        if top1_score >= self.resolver_high_confidence and (top1_score - top2_score) >= self.resolver_high_margin:
            return {
                "status": "resolved",
                "query_text": brand_query,
                "match": {
                    "brand_id": top1_id,
                    "brand_name": self.master_by_id[top1_id]["brand_name"],
                    "confidence": round(top1_score, 4),
                },
                "candidates": candidates or raw_candidates[:1],
                "reason": "fuzzy_confident",
            }
        if top1_score >= self.resolver_ambiguous_min:
            return {
                "status": "ambiguous",
                "query_text": brand_query,
                "match": None,
                "candidates": candidates or raw_candidates,
                "reason": "fuzzy_ambiguous",
            }

        # Optional semantic fallback (no hardcoded alias): use LLM only when API key exists.
        llm_pick = self._llm_resolve_brand(brand_query)
        if llm_pick and llm_pick["confidence"] >= self.resolver_llm_min:
            return {
                "status": "resolved",
                "query_text": brand_query,
                "match": {
                    "brand_id": llm_pick["brand_id"],
                    "brand_name": llm_pick["brand_name"],
                    "confidence": round(llm_pick["confidence"], 4),
                },
                "candidates": [
                    {
                        "brand_id": llm_pick["brand_id"],
                        "brand_name": llm_pick["brand_name"],
                        "confidence": round(llm_pick["confidence"], 4),
                        "stage": "llm_fallback",
                    }
                ],
                "reason": f"llm_fallback:{llm_pick['reason']}",
            }
        return {
            "status": "not_found",
            "query_text": brand_query,
            "match": None,
            "candidates": candidates,
            "reason": "low_confidence",
        }

    def resolve_brand_debug(self, brand_query: str, top_k: int = 3) -> dict[str, Any]:
        """Return raw resolver decision payload for diagnostics/tuning."""
        return self._resolve_brand_candidate(brand_query=brand_query, top_k=top_k)

    def resolve_brand(self, brand_query: str) -> dict[str, Any]:
        res = self._resolve_brand_candidate(brand_query, top_k=3)
        if res["status"] == "resolved" and res["match"]:
            return self.master_by_id[res["match"]["brand_id"]]
        if res["status"] == "ambiguous":
            raise BrandResolutionError(
                query_text=brand_query,
                status="ambiguous",
                candidates=res.get("candidates", []),
                reason=res.get("reason"),
            )
        raise BrandResolutionError(
            query_text=brand_query,
            status="not_found",
            candidates=res.get("candidates", []),
            reason=res.get("reason"),
        )

    def _pick_richer_row(self, left: dict[str, Any] | None, right: dict[str, Any]) -> dict[str, Any]:
        if left is None:
            return right
        left_score = sum(1 for v in left.values() if v is not None and str(v).strip() != "")
        right_score = sum(1 for v in right.values() if v is not None and str(v).strip() != "")
        if right_score >= left_score:
            return right
        return left

    def _upsert_year_row(
        self,
        table: dict[int, dict[int, dict[str, Any]]],
        brand_id: int,
        year: int | None,
        row: dict[str, Any],
    ) -> None:
        if year is None:
            return
        existing = table[brand_id].get(year)
        table[brand_id][year] = self._pick_richer_row(existing, row)

    def _load_api_data_enrichment(self) -> None:
        brand_list_rows = self.selected_source.load_brand_list_info()
        frcs_rows = self.selected_source.load_brand_frcs_stats()
        fntn_rows = self.selected_source.load_brand_fntn_stats()
        brand_rows = self.selected_source.load_brand_brand_stats()
        interior_rows = self.selected_source.load_brand_interior_cost()

        # Build fallback id maps even in build mode, so enrichment can map selected rows.
        if not self._api_mnno_to_id and not self._api_name_to_id:
            next_id = 1
            for row in brand_list_rows:
                mnno = str(row.get("brandMnno") or "").strip()
                name = str(row.get("brandNm") or "").strip()
                existing = None
                if name:
                    for bid, master in self.master_by_id.items():
                        if str(master.get("brand_name") or "").strip() == name:
                            existing = bid
                            break
                if existing is None:
                    if mnno and mnno in self._api_mnno_to_id:
                        existing = self._api_mnno_to_id[mnno]
                    elif name and name in self._api_name_to_id:
                        existing = self._api_name_to_id[name]
                    else:
                        existing = next_id
                        next_id += 1
                if mnno and mnno not in self._api_mnno_to_id:
                    self._api_mnno_to_id[mnno] = existing
                if name and name not in self._api_name_to_id:
                    self._api_name_to_id[name] = existing

        for row in brand_list_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None or brand_id not in self.master_by_id:
                continue
            year = _safe_int(row.get("jngBizCrtraYr"))
            payload = {
                "brand_mnno": row.get("brandMnno"),
                "hq_mnno": row.get("jnghdqrtrsMnno"),
                "business_registration_no": row.get("brno"),
                "corporate_registration_no": row.get("crno"),
                "representative_name": row.get("jnghdqrtrsRprsvNm"),
                "major_products": row.get("majrGdsNm"),
                "franchise_start_date_source": row.get("jngBizStrtDate"),
                "year": year,
            }
            current = self.brand_list_info_latest_by_brand.get(brand_id)
            current_year = _safe_int((current or {}).get("year"))
            if current is None or year is None or current_year is None or year >= current_year:
                self.brand_list_info_latest_by_brand[brand_id] = payload

        for row in frcs_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None or brand_id not in self.master_by_id:
                continue
            year = _safe_int(row.get("yr"))
            payload = {
                "contract_end_count": _safe_int(row.get("ctrtEndCnt")),
                "contract_cancel_count": _safe_int(row.get("ctrtCncltnCnt")),
                "name_change_count": _safe_int(row.get("nmChgCnt")),
                "new_store_registrations": _safe_int(row.get("newFrcsRgsCnt")),
                "avg_sales_per_area_krw": _safe_int(row.get("arUnitAvrgSlsAmt")),
                "avg_sales_total_krw": _safe_int(row.get("avrgSlsAmt")),
            }
            self._upsert_year_row(self.frcs_extra_by_brand_year, brand_id, year, payload)

        for row in fntn_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None or brand_id not in self.master_by_id:
                continue
            year = _safe_int(row.get("yr"))
            payload = {
                "startup_deposit_krw": _safe_int(row.get("jngBzmnJngAmt")),
                "startup_training_krw": _safe_int(row.get("jngBzmnEduAmt")),
                "startup_other_krw": _safe_int(row.get("jngBzmnEtcAmt")),
                "startup_guarantee_krw": _safe_int(row.get("jngBzmnAssrncAmt")),
                "startup_sum_krw": _safe_int(row.get("smtnAmt")),
            }
            self._upsert_year_row(self.fntn_extra_by_brand_year, brand_id, year, payload)

        for row in brand_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None or brand_id not in self.master_by_id:
                continue
            year = _safe_int(row.get("yr"))
            payload = {
                "executives_count": _safe_int(row.get("allExctvCnt")),
                "employees_count": _safe_int(row.get("empCnt")),
                "franchise_business_years_text": row.get("jngBizYycnt"),
                "franchise_start_date_source": row.get("jngBizStrtDate"),
            }
            self._upsert_year_row(self.brand_extra_by_brand_year, brand_id, year, payload)

        for row in interior_rows:
            brand_id = self._resolve_api_brand_id(row)
            if brand_id is None or brand_id not in self.master_by_id:
                continue
            year = _safe_int(row.get("jngBizCrtraYr"))
            payload = {
                "interior_currency": row.get("crrncyUnitCdNm"),
                "interior_store_area": _safe_int(row.get("storCrtraAr")),
                "interior_cost_range_text": row.get("intrrAmtScopeVal"),
                "interior_cost_per_area_range_text": row.get("unitArIntrrAmtScopeVal"),
                "interior_cost_mid_krw": _parse_range_mid(row.get("intrrAmtScopeVal")),
                "interior_cost_per_area_mid_krw": _parse_range_mid(row.get("unitArIntrrAmtScopeVal")),
            }
            self._upsert_year_row(self.interior_extra_by_brand_year, brand_id, year, payload)

    def _pick_extra_by_year(
        self,
        year_map: dict[int, dict[int, dict[str, Any]]],
        brand_id: int,
        preferred_year: int | None,
    ) -> tuple[dict[str, Any] | None, int | None]:
        rows = year_map.get(brand_id, {})
        if not rows:
            return None, None
        years = sorted(rows.keys())
        if preferred_year is None:
            y = years[-1]
            return rows[y], y
        if preferred_year in rows:
            return rows[preferred_year], preferred_year
        prior = [y for y in years if y <= preferred_year]
        if prior:
            y = prior[-1]
            return rows[y], y
        y = years[0]
        return rows[y], y

    def _extended_context(self, brand_id: int, preferred_year: int | None) -> dict[str, Any]:
        info = self.brand_list_info_latest_by_brand.get(brand_id, {})
        frcs, frcs_year = self._pick_extra_by_year(self.frcs_extra_by_brand_year, brand_id, preferred_year)
        fntn, fntn_year = self._pick_extra_by_year(self.fntn_extra_by_brand_year, brand_id, preferred_year)
        brand, brand_year = self._pick_extra_by_year(self.brand_extra_by_brand_year, brand_id, preferred_year)
        interior, interior_year = self._pick_extra_by_year(
            self.interior_extra_by_brand_year, brand_id, preferred_year
        )
        return {
            "identity": info or {},
            "franchise_ops": {"year_used": frcs_year, **(frcs or {})},
            "funding": {"year_used": fntn_year, **(fntn or {})},
            "organization": {"year_used": brand_year, **(brand or {})},
            "interior": {"year_used": interior_year, **(interior or {})},
        }

    def _build_metrics_for_brand_year(
        self,
        brand_id: int,
        year_row: dict[str, Any],
        cost_summary: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        year_used = year_row["year"]
        frcs, _ = self._pick_extra_by_year(self.frcs_extra_by_brand_year, brand_id, year_used)
        fntn, _ = self._pick_extra_by_year(self.fntn_extra_by_brand_year, brand_id, year_used)
        brand, _ = self._pick_extra_by_year(self.brand_extra_by_brand_year, brand_id, year_used)
        interior, _ = self._pick_extra_by_year(self.interior_extra_by_brand_year, brand_id, year_used)

        return {
            "store_count": year_row["store_count"],
            "new_stores": year_row["new_stores"],
            "closed_stores": year_row["closed_stores"],
            "net_store_change": year_row["net_store_change"],
            "store_growth_rate": year_row["store_growth_rate"],
            "closure_rate": year_row["closure_rate"],
            "churn_rate": year_row["churn_rate"],
            "avg_sales_krw": year_row.get("avg_sales_krw"),
            "startup_total_initial_cost_krw": None
            if cost_summary is None
            else cost_summary.get("total_initial_cost_krw"),
            "contract_end_count": (frcs or {}).get("contract_end_count"),
            "contract_cancel_count": (frcs or {}).get("contract_cancel_count"),
            "name_change_count": (frcs or {}).get("name_change_count"),
            "new_store_registrations": (frcs or {}).get("new_store_registrations"),
            "avg_sales_per_area_krw": (frcs or {}).get("avg_sales_per_area_krw"),
            "avg_sales_total_krw": (frcs or {}).get("avg_sales_total_krw"),
            "startup_deposit_krw": (fntn or {}).get("startup_deposit_krw"),
            "startup_training_krw": (fntn or {}).get("startup_training_krw"),
            "startup_other_krw": (fntn or {}).get("startup_other_krw"),
            "startup_guarantee_krw": (fntn or {}).get("startup_guarantee_krw"),
            "startup_sum_krw": (fntn or {}).get("startup_sum_krw"),
            "executives_count": (brand or {}).get("executives_count"),
            "employees_count": (brand or {}).get("employees_count"),
            "interior_store_area": (interior or {}).get("interior_store_area"),
            "interior_cost_mid_krw": (interior or {}).get("interior_cost_mid_krw"),
            "interior_cost_per_area_mid_krw": (interior or {}).get("interior_cost_per_area_mid_krw"),
        }

    def _pick_year_row(self, rows: list[dict[str, Any]], year: int | None) -> dict[str, Any]:
        if not rows:
            raise ValueError("해당 브랜드의 연도별 통계 데이터가 없습니다.")

        if year is None:
            return max(rows, key=lambda r: r["year"])

        exact = [r for r in rows if r["year"] == year]
        if exact:
            return exact[0]

        prior = [r for r in rows if r["year"] < year]
        if prior:
            return max(prior, key=lambda r: r["year"])

        raise ValueError(f"요청한 연도({year}) 또는 그 이전 연도의 통계 데이터가 없습니다.")

    def _pick_common_year(self, brand_id_a: int, brand_id_b: int, year: int | None) -> int:
        years_a = {r["year"] for r in self.year_stats_by_brand.get(brand_id_a, [])}
        years_b = {r["year"] for r in self.year_stats_by_brand.get(brand_id_b, [])}
        common = sorted(years_a & years_b)
        if not common:
            raise ValueError("두 브랜드에 공통으로 존재하는 연도가 없습니다.")

        if year is None:
            return max(common)

        if year in common:
            return year

        prior = [y for y in common if y < year]
        if prior:
            return max(prior)

        raise ValueError(f"요청한 연도({year}) 또는 그 이전 공통 연도의 통계 데이터가 없습니다.")

    def _pick_exact_year_row(self, rows: list[dict[str, Any]], year: int) -> dict[str, Any]:
        for row in rows:
            if row["year"] == year:
                return row
        raise ValueError(f"{year}년의 정확한 통계 행이 없습니다.")

    def _select_cost_summary(
        self, brand_id: int, preferred_year: int, requested_store_type: str | None
    ) -> dict[str, Any]:
        candidate_keys = [k for k in self.costs_by_brand_year_type if k[0] == brand_id]
        if not candidate_keys:
            return {
                "year_used": None,
                "store_type_used": None,
                "cost_basis": "no_cost_data",
                "total_initial_cost_krw": None,
                "cost_breakdown_krw": {},
            }

        years = sorted({k[1] for k in candidate_keys})
        year_candidates = [y for y in years if y <= preferred_year] or years
        year_used = max(year_candidates)

        year_keys = [k for k in candidate_keys if k[1] == year_used]
        types = sorted({k[2] for k in year_keys})

        def totals_for_type(store_type: str) -> tuple[int | None, dict[str, int], str]:
            categories = self.costs_by_brand_year_type[(brand_id, year_used, store_type)]
            if "total_initial_cost" in categories:
                return categories["total_initial_cost"], categories, "reported_total_initial_cost"
            computed = sum(categories.values())
            return computed, categories, "computed_from_components"

        selected_type: str | None = None
        if requested_store_type:
            low = requested_store_type.lower()
            for st in types:
                if st.lower() == low:
                    selected_type = st
                    break
            if selected_type is None:
                return {
                    "year_used": year_used,
                    "store_type_used": None,
                    "available_store_types": types,
                    "cost_basis": "requested_store_type_not_found",
                    "total_initial_cost_krw": None,
                    "cost_breakdown_krw": {},
                }
        else:
            for st in types:
                if st.lower() == "standard":
                    selected_type = st
                    break
            if selected_type is None:
                best_type = None
                best_total = None
                best_basis = ""
                for st in types:
                    total, _, basis = totals_for_type(st)
                    if total is None:
                        continue
                    if best_total is None or total < best_total:
                        best_total = total
                        best_type = st
                        best_basis = basis
                if best_type is None:
                    return {
                        "year_used": year_used,
                        "store_type_used": None,
                        "cost_basis": "no_usable_cost_data",
                        "total_initial_cost_krw": None,
                        "cost_breakdown_krw": {},
                    }
                selected_type = best_type

        total, categories, basis = totals_for_type(selected_type)
        return {
            "year_used": year_used,
            "store_type_used": selected_type,
            "available_store_types": types,
            "cost_basis": basis,
            "total_initial_cost_krw": total,
            "cost_breakdown_krw": categories,
        }

    def get_brand_overview(
        self, brand_name: str, year: int | None = None, store_type: str | None = None
    ) -> dict[str, Any]:
        brand = self.resolve_brand(brand_name)
        brand_id = brand["brand_id"]

        stats_rows = self.year_stats_by_brand.get(brand_id, [])
        if not stats_rows:
            cost_summary = self._select_cost_summary(brand_id, preferred_year=None, requested_store_type=store_type)
            extended = self._extended_context(brand_id=brand_id, preferred_year=None)
            store_types = [
                {
                    "store_type": row["store_type"],
                    "standard_area_pyeong": row["standard_area_pyeong"],
                }
                for row in sorted(self.store_types_by_brand.get(brand_id, []), key=lambda r: r["store_type"])
            ]
            return {
                "brand": {
                    "brand_id": brand_id,
                    "brand_name": brand["brand_name"],
                    "company_name": brand.get("company_name"),
                    "franchise_start_date": brand.get("franchise_start_date"),
                    "category_main": brand.get("category_main"),
                    "category_sub": brand.get("category_sub"),
                },
                "year_used": None,
                "stats": {
                    "store_count": None,
                    "new_stores": None,
                    "closed_stores": None,
                    "net_store_change": None,
                    "store_growth_rate": None,
                    "closure_rate": None,
                    "churn_rate": None,
                    "avg_sales_krw": None,
                },
                "store_types": store_types,
                "startup_cost": cost_summary,
                "extended": extended,
                "formatted": {
                    "store_count": "N/A",
                    "new_stores": "N/A",
                    "closed_stores": "N/A",
                    "net_store_change": "N/A",
                    "store_growth_rate": "N/A",
                    "closure_rate": "N/A",
                    "churn_rate": "N/A",
                    "avg_sales_krw": "not disclosed in source data",
                    "startup_total_initial_cost_krw": _format_krw(cost_summary["total_initial_cost_krw"]),
                },
                "note": "연도별 통계 데이터가 없어 브랜드 기본 정보만 제공합니다.",
            }

        year_row = self._pick_year_row(stats_rows, year)
        year_used = year_row["year"]
        cost_summary = self._select_cost_summary(brand_id, preferred_year=year_used, requested_store_type=store_type)
        all_metrics = self._build_metrics_for_brand_year(
            brand_id=brand_id,
            year_row=year_row,
            cost_summary=cost_summary,
        )
        extended = self._extended_context(brand_id=brand_id, preferred_year=year_used)

        store_types = [
            {
                "store_type": row["store_type"],
                "standard_area_pyeong": row["standard_area_pyeong"],
            }
            for row in sorted(self.store_types_by_brand.get(brand_id, []), key=lambda r: r["store_type"])
        ]

        avg_sales = year_row.get("avg_sales_krw")
        result = {
            "brand": {
                "brand_id": brand_id,
                "brand_name": brand["brand_name"],
                "company_name": brand["company_name"],
                "franchise_start_date": brand["franchise_start_date"],
                "category_main": brand["category_main"],
                "category_sub": brand["category_sub"],
            },
            "year_used": year_used,
            "stats": {
                "store_count": year_row["store_count"],
                "new_stores": year_row["new_stores"],
                "closed_stores": year_row["closed_stores"],
                "net_store_change": year_row["net_store_change"],
                "store_growth_rate": year_row["store_growth_rate"],
                "closure_rate": year_row["closure_rate"],
                "churn_rate": year_row["churn_rate"],
                "avg_sales_krw": avg_sales,
            },
            "all_metrics": all_metrics,
            "store_types": store_types,
            "startup_cost": cost_summary,
            "extended": extended,
            "formatted": {
                "store_count": _format_int(year_row["store_count"]),
                "new_stores": _format_int(year_row["new_stores"]),
                "closed_stores": _format_int(year_row["closed_stores"]),
                "net_store_change": _format_int(year_row["net_store_change"]),
                "store_growth_rate": _format_pct(year_row["store_growth_rate"]),
                "closure_rate": _format_pct(year_row["closure_rate"]),
                "churn_rate": _format_pct(year_row["churn_rate"]),
                "avg_sales_krw": _format_krw(avg_sales),
                "startup_total_initial_cost_krw": _format_krw(cost_summary["total_initial_cost_krw"]),
            },
        }
        return result

    def get_brand_compare(
        self,
        brand_a_name: str,
        brand_b_name: str,
        year: int | None = None,
        store_type: str | None = None,
    ) -> dict[str, Any]:
        brand_a = self.resolve_brand(brand_a_name)
        brand_b = self.resolve_brand(brand_b_name)
        if brand_a["brand_id"] == brand_b["brand_id"]:
            raise ValueError("비교를 위해 서로 다른 두 브랜드를 입력해 주세요.")

        year_used = self._pick_common_year(brand_a["brand_id"], brand_b["brand_id"], year)
        row_a = self._pick_exact_year_row(self.year_stats_by_brand[brand_a["brand_id"]], year_used)
        row_b = self._pick_exact_year_row(self.year_stats_by_brand[brand_b["brand_id"]], year_used)

        cost_a = self._select_cost_summary(
            brand_id=brand_a["brand_id"], preferred_year=year_used, requested_store_type=store_type
        )
        cost_b = self._select_cost_summary(
            brand_id=brand_b["brand_id"], preferred_year=year_used, requested_store_type=store_type
        )
        metrics_a = self._build_metrics_for_brand_year(
            brand_id=brand_a["brand_id"],
            year_row=row_a,
            cost_summary=cost_a,
        )
        metrics_b = self._build_metrics_for_brand_year(
            brand_id=brand_b["brand_id"],
            year_row=row_b,
            cost_summary=cost_b,
        )

        def pack(
            brand: dict[str, Any],
            row: dict[str, Any],
            cost: dict[str, Any],
            all_metrics: dict[str, Any],
        ) -> dict[str, Any]:
            avg_sales = row.get("avg_sales_krw")
            return {
                "brand": {
                    "brand_id": brand["brand_id"],
                    "brand_name": brand["brand_name"],
                    "company_name": brand["company_name"],
                },
                "year_used": year_used,
                "stats": {
                    "store_count": row["store_count"],
                    "new_stores": row["new_stores"],
                    "closed_stores": row["closed_stores"],
                    "net_store_change": row["net_store_change"],
                    "store_growth_rate": row["store_growth_rate"],
                    "closure_rate": row["closure_rate"],
                    "churn_rate": row["churn_rate"],
                    "avg_sales_krw": avg_sales,
                },
                "all_metrics": all_metrics,
                "extended": self._extended_context(brand["brand_id"], preferred_year=year_used),
                "startup_cost": {
                    "year_used": cost["year_used"],
                    "store_type_used": cost["store_type_used"],
                    "cost_basis": cost["cost_basis"],
                    "total_initial_cost_krw": cost["total_initial_cost_krw"],
                },
                "formatted": {
                    "store_count": _format_int(row["store_count"]),
                    "new_stores": _format_int(row["new_stores"]),
                    "closed_stores": _format_int(row["closed_stores"]),
                    "net_store_change": _format_int(row["net_store_change"]),
                    "store_growth_rate": _format_pct(row["store_growth_rate"]),
                    "closure_rate": _format_pct(row["closure_rate"]),
                    "churn_rate": _format_pct(row["churn_rate"]),
                    "avg_sales_krw": _format_krw(avg_sales),
                    "startup_total_initial_cost_krw": _format_krw(cost["total_initial_cost_krw"]),
                },
            }

        cost_diff = None
        if (
            cost_a.get("total_initial_cost_krw") is not None
            and cost_b.get("total_initial_cost_krw") is not None
        ):
            cost_diff = cost_a["total_initial_cost_krw"] - cost_b["total_initial_cost_krw"]

        return {
            "comparison_year_used": year_used,
            "store_type_requested": store_type,
            "brand_a": pack(brand_a, row_a, cost_a, metrics_a),
            "brand_b": pack(brand_b, row_b, cost_b, metrics_b),
            "diff": {
                "store_count": row_a["store_count"] - row_b["store_count"],
                "new_stores": row_a["new_stores"] - row_b["new_stores"],
                "closed_stores": row_a["closed_stores"] - row_b["closed_stores"],
                "net_store_change": row_a["net_store_change"] - row_b["net_store_change"],
                "store_growth_rate": row_a["store_growth_rate"] - row_b["store_growth_rate"],
                "closure_rate": row_a["closure_rate"] - row_b["closure_rate"],
                "churn_rate": row_a["churn_rate"] - row_b["churn_rate"],
                "avg_sales_krw": (
                    None
                    if row_a.get("avg_sales_krw") is None or row_b.get("avg_sales_krw") is None
                    else row_a["avg_sales_krw"] - row_b["avg_sales_krw"]
                ),
                "startup_total_initial_cost_krw": cost_diff,
                "contract_end_count": (
                    None
                    if metrics_a.get("contract_end_count") is None or metrics_b.get("contract_end_count") is None
                    else metrics_a["contract_end_count"] - metrics_b["contract_end_count"]
                ),
                "contract_cancel_count": (
                    None
                    if metrics_a.get("contract_cancel_count") is None or metrics_b.get("contract_cancel_count") is None
                    else metrics_a["contract_cancel_count"] - metrics_b["contract_cancel_count"]
                ),
                "name_change_count": (
                    None
                    if metrics_a.get("name_change_count") is None or metrics_b.get("name_change_count") is None
                    else metrics_a["name_change_count"] - metrics_b["name_change_count"]
                ),
                "new_store_registrations": (
                    None
                    if metrics_a.get("new_store_registrations") is None
                    or metrics_b.get("new_store_registrations") is None
                    else metrics_a["new_store_registrations"] - metrics_b["new_store_registrations"]
                ),
                "avg_sales_per_area_krw": (
                    None
                    if metrics_a.get("avg_sales_per_area_krw") is None
                    or metrics_b.get("avg_sales_per_area_krw") is None
                    else metrics_a["avg_sales_per_area_krw"] - metrics_b["avg_sales_per_area_krw"]
                ),
                "startup_sum_krw": (
                    None
                    if metrics_a.get("startup_sum_krw") is None or metrics_b.get("startup_sum_krw") is None
                    else metrics_a["startup_sum_krw"] - metrics_b["startup_sum_krw"]
                ),
                "executives_count": (
                    None
                    if metrics_a.get("executives_count") is None or metrics_b.get("executives_count") is None
                    else metrics_a["executives_count"] - metrics_b["executives_count"]
                ),
                "employees_count": (
                    None
                    if metrics_a.get("employees_count") is None or metrics_b.get("employees_count") is None
                    else metrics_a["employees_count"] - metrics_b["employees_count"]
                ),
                "interior_cost_mid_krw": (
                    None
                    if metrics_a.get("interior_cost_mid_krw") is None or metrics_b.get("interior_cost_mid_krw") is None
                    else metrics_a["interior_cost_mid_krw"] - metrics_b["interior_cost_mid_krw"]
                ),
            },
            "diff_formatted": {
                "store_count": _format_int(row_a["store_count"] - row_b["store_count"]),
                "new_stores": _format_int(row_a["new_stores"] - row_b["new_stores"]),
                "closed_stores": _format_int(row_a["closed_stores"] - row_b["closed_stores"]),
                "net_store_change": _format_int(row_a["net_store_change"] - row_b["net_store_change"]),
                "store_growth_rate": _format_pct(row_a["store_growth_rate"] - row_b["store_growth_rate"]),
                "closure_rate": _format_pct(row_a["closure_rate"] - row_b["closure_rate"]),
                "churn_rate": _format_pct(row_a["churn_rate"] - row_b["churn_rate"]),
                "avg_sales_krw": _format_krw(
                    None
                    if row_a.get("avg_sales_krw") is None or row_b.get("avg_sales_krw") is None
                    else row_a["avg_sales_krw"] - row_b["avg_sales_krw"]
                ),
                "startup_total_initial_cost_krw": _format_krw(cost_diff),
                "contract_end_count": _format_int(
                    None
                    if metrics_a.get("contract_end_count") is None or metrics_b.get("contract_end_count") is None
                    else metrics_a["contract_end_count"] - metrics_b["contract_end_count"]
                ),
                "contract_cancel_count": _format_int(
                    None
                    if metrics_a.get("contract_cancel_count") is None or metrics_b.get("contract_cancel_count") is None
                    else metrics_a["contract_cancel_count"] - metrics_b["contract_cancel_count"]
                ),
                "name_change_count": _format_int(
                    None
                    if metrics_a.get("name_change_count") is None or metrics_b.get("name_change_count") is None
                    else metrics_a["name_change_count"] - metrics_b["name_change_count"]
                ),
                "new_store_registrations": _format_int(
                    None
                    if metrics_a.get("new_store_registrations") is None
                    or metrics_b.get("new_store_registrations") is None
                    else metrics_a["new_store_registrations"] - metrics_b["new_store_registrations"]
                ),
                "avg_sales_per_area_krw": _format_krw(
                    None
                    if metrics_a.get("avg_sales_per_area_krw") is None
                    or metrics_b.get("avg_sales_per_area_krw") is None
                    else metrics_a["avg_sales_per_area_krw"] - metrics_b["avg_sales_per_area_krw"]
                ),
                "startup_sum_krw": _format_krw(
                    None
                    if metrics_a.get("startup_sum_krw") is None or metrics_b.get("startup_sum_krw") is None
                    else metrics_a["startup_sum_krw"] - metrics_b["startup_sum_krw"]
                ),
                "executives_count": _format_int(
                    None
                    if metrics_a.get("executives_count") is None or metrics_b.get("executives_count") is None
                    else metrics_a["executives_count"] - metrics_b["executives_count"]
                ),
                "employees_count": _format_int(
                    None
                    if metrics_a.get("employees_count") is None or metrics_b.get("employees_count") is None
                    else metrics_a["employees_count"] - metrics_b["employees_count"]
                ),
                "interior_cost_mid_krw": _format_krw(
                    None
                    if metrics_a.get("interior_cost_mid_krw") is None or metrics_b.get("interior_cost_mid_krw") is None
                    else metrics_a["interior_cost_mid_krw"] - metrics_b["interior_cost_mid_krw"]
                ),
            },
        }

    def _evaluate_condition(self, left: Any, op: str, right: Any) -> bool:
        if left is None:
            return False
        if op == "<":
            return left < right
        if op == "<=":
            return left <= right
        if op == ">":
            return left > right
        if op == ">=":
            return left >= right
        if op == "==":
            return left == right
        if op == "!=":
            return left != right
        raise ValueError(f"지원하지 않는 연산자입니다: '{op}'. 사용 가능: <, <=, >, >=, ==, !=")

    def _apply_sort_specs(
        self,
        items: list[dict[str, Any]],
        sort_specs: list[dict[str, str]],
    ) -> list[dict[str, Any]]:
        out = list(items)
        for spec in reversed(sort_specs):
            field = spec["field"]
            order = spec["order"]
            if order not in {"asc", "desc"}:
                raise ValueError(f"지원하지 않는 정렬 순서입니다: '{order}'. 'asc' 또는 'desc'를 사용하세요.")

            def key_fn(item: dict[str, Any]) -> tuple[bool, float]:
                val = item["metrics"].get(field)
                if val is None:
                    return (True, float("inf"))
                num = float(val)
                return (False, num if order == "asc" else -num)

            out.sort(key=key_fn)
        return out

    def get_brand_filter_search(
        self,
        conditions: list[dict[str, Any]],
        year: int | None = None,
        store_type: str | None = None,
        sort_by: str | None = None,
        sort_order: str = "asc",
        limit: int = 10,
    ) -> dict[str, Any]:
        supported_fields = {
            "store_count",
            "new_stores",
            "closed_stores",
            "net_store_change",
            "store_growth_rate",
            "closure_rate",
            "churn_rate",
            "avg_sales_krw",
            "startup_total_initial_cost_krw",
            "contract_end_count",
            "contract_cancel_count",
            "name_change_count",
            "new_store_registrations",
            "avg_sales_per_area_krw",
            "avg_sales_total_krw",
            "startup_deposit_krw",
            "startup_training_krw",
            "startup_other_krw",
            "startup_guarantee_krw",
            "startup_sum_krw",
            "executives_count",
            "employees_count",
            "interior_store_area",
            "interior_cost_mid_krw",
            "interior_cost_per_area_mid_krw",
        }

        if limit <= 0:
            raise ValueError("limit 값은 1 이상이어야 합니다.")

        normalized_conditions = conditions or []
        for cond in normalized_conditions:
            field = cond.get("field")
            op = cond.get("op")
            if field not in supported_fields:
                raise ValueError(f"지원하지 않는 필터 필드입니다: '{field}'.")
            if op not in {"<", "<=", ">", ">=", "==", "!="}:
                raise ValueError(f"지원하지 않는 연산자입니다: '{op}'.")

        candidates: list[dict[str, Any]] = []
        for brand in self.brand_master:
            brand_id = brand["brand_id"]
            stats_rows = self.year_stats_by_brand.get(brand_id, [])
            if not stats_rows:
                continue

            year_row = self._pick_year_row(stats_rows, year)
            cost = self._select_cost_summary(
                brand_id=brand_id,
                preferred_year=year_row["year"],
                requested_store_type=store_type,
            )

            metrics = self._build_metrics_for_brand_year(
                brand_id=brand_id,
                year_row=year_row,
                cost_summary=cost,
            )

            passed = True
            for cond in normalized_conditions:
                field = cond["field"]
                op = cond["op"]
                val = cond["value"]
                if not self._evaluate_condition(metrics.get(field), op, val):
                    passed = False
                    break
            if not passed:
                continue

            candidates.append(
                {
                    "brand": {
                        "brand_id": brand_id,
                        "brand_name": brand["brand_name"],
                        "company_name": brand["company_name"],
                    },
                    "year_used": year_row["year"],
                    "store_type_used_for_cost": cost.get("store_type_used"),
                    "metrics": metrics,
                    "extended": self._extended_context(brand_id=brand_id, preferred_year=year_row["year"]),
                    "formatted": {
                        "store_count": _format_int(metrics["store_count"]),
                        "new_stores": _format_int(metrics["new_stores"]),
                        "closed_stores": _format_int(metrics["closed_stores"]),
                        "net_store_change": _format_int(metrics["net_store_change"]),
                        "store_growth_rate": _format_pct(metrics["store_growth_rate"]),
                        "closure_rate": _format_pct(metrics["closure_rate"]),
                        "churn_rate": _format_pct(metrics["churn_rate"]),
                        "avg_sales_krw": _format_krw(metrics["avg_sales_krw"]),
                        "startup_total_initial_cost_krw": _format_krw(metrics["startup_total_initial_cost_krw"]),
                        "contract_end_count": _format_int(metrics["contract_end_count"]),
                        "contract_cancel_count": _format_int(metrics["contract_cancel_count"]),
                        "name_change_count": _format_int(metrics["name_change_count"]),
                        "new_store_registrations": _format_int(metrics["new_store_registrations"]),
                        "avg_sales_per_area_krw": _format_krw(metrics["avg_sales_per_area_krw"]),
                        "avg_sales_total_krw": _format_krw(metrics["avg_sales_total_krw"]),
                        "startup_deposit_krw": _format_krw(metrics["startup_deposit_krw"]),
                        "startup_training_krw": _format_krw(metrics["startup_training_krw"]),
                        "startup_other_krw": _format_krw(metrics["startup_other_krw"]),
                        "startup_guarantee_krw": _format_krw(metrics["startup_guarantee_krw"]),
                        "startup_sum_krw": _format_krw(metrics["startup_sum_krw"]),
                        "executives_count": _format_int(metrics["executives_count"]),
                        "employees_count": _format_int(metrics["employees_count"]),
                        "interior_store_area": _format_int(metrics["interior_store_area"]),
                        "interior_cost_mid_krw": _format_krw(metrics["interior_cost_mid_krw"]),
                        "interior_cost_per_area_mid_krw": _format_krw(
                            metrics["interior_cost_per_area_mid_krw"]
                        ),
                    },
                }
            )

        if sort_by:
            if sort_by not in supported_fields:
                raise ValueError(f"지원하지 않는 정렬 필드입니다: '{sort_by}'.")
            sort_specs = [{"field": sort_by, "order": sort_order}]
        else:
            sort_specs = [
                {"field": "churn_rate", "order": "asc"},
                {"field": "store_growth_rate", "order": "desc"},
                {"field": "store_count", "order": "desc"},
            ]

        sorted_candidates = self._apply_sort_specs(candidates, sort_specs)
        selected = sorted_candidates[:limit]

        return {
            "filters_applied": normalized_conditions,
            "year_requested": year,
            "store_type_requested": store_type,
            "sort_applied": sort_specs,
            "limit": limit,
            "total_matches": len(sorted_candidates),
            "results": selected,
        }

    def get_brand_trend(
        self,
        brand_name: str,
        start_year: int | None = None,
        end_year: int | None = None,
        metrics: list[str] | None = None,
    ) -> dict[str, Any]:
        supported_metrics = [
            "store_count",
            "new_stores",
            "closed_stores",
            "net_store_change",
            "store_growth_rate",
            "closure_rate",
            "churn_rate",
            "avg_sales_krw",
            "contract_end_count",
            "contract_cancel_count",
            "name_change_count",
            "new_store_registrations",
            "avg_sales_per_area_krw",
            "avg_sales_total_krw",
            "startup_deposit_krw",
            "startup_training_krw",
            "startup_other_krw",
            "startup_guarantee_krw",
            "startup_sum_krw",
            "executives_count",
            "employees_count",
            "interior_store_area",
            "interior_cost_mid_krw",
            "interior_cost_per_area_mid_krw",
        ]
        metric_set = set(metrics) if metrics else set(supported_metrics)
        unsupported = [m for m in metric_set if m not in supported_metrics]
        if unsupported:
            raise ValueError(f"지원하지 않는 추이 지표입니다: {unsupported}")

        brand = self.resolve_brand(brand_name)
        rows = self.year_stats_by_brand.get(brand["brand_id"], [])
        if not rows:
            raise ValueError("해당 브랜드의 연도별 통계 데이터가 없습니다.")

        lo = start_year if start_year is not None else min(r["year"] for r in rows)
        hi = end_year if end_year is not None else max(r["year"] for r in rows)
        if lo > hi:
            raise ValueError("start_year는 end_year보다 작거나 같아야 합니다.")

        timeline_rows = [r for r in rows if lo <= r["year"] <= hi]
        if not timeline_rows:
            raise ValueError("요청한 연도 구간에 데이터가 없습니다.")

        timeline = []
        merged_by_year: dict[int, dict[str, Any]] = {}
        for row in timeline_rows:
            point_raw = {"year": row["year"]}
            point_fmt = {"year": row["year"]}
            merged_metrics = self._build_metrics_for_brand_year(
                brand_id=brand["brand_id"],
                year_row=row,
                cost_summary=None,
            )
            merged_by_year[row["year"]] = merged_metrics
            for m in metric_set:
                val = merged_metrics.get(m)
                point_raw[m] = val
                if m in {"store_growth_rate", "closure_rate", "churn_rate"}:
                    point_fmt[m] = _format_pct(val)
                elif m == "avg_sales_krw" or m.endswith("_krw"):
                    point_fmt[m] = _format_krw(val)
                else:
                    point_fmt[m] = _format_int(val)
            timeline.append({"raw": point_raw, "formatted": point_fmt})

        first = timeline_rows[0]
        last = timeline_rows[-1]
        first_metrics = merged_by_year[first["year"]]
        last_metrics = merged_by_year[last["year"]]
        summary: dict[str, Any] = {}
        for m in metric_set:
            start_val = first_metrics.get(m)
            end_val = last_metrics.get(m)
            delta = None if start_val is None or end_val is None else end_val - start_val
            if delta is None:
                trend = "insufficient_data"
            elif delta > 0:
                trend = "up"
            elif delta < 0:
                trend = "down"
            else:
                trend = "flat"
            if m in {"store_growth_rate", "closure_rate", "churn_rate"}:
                fmt_start = _format_pct(start_val)
                fmt_end = _format_pct(end_val)
                fmt_delta = _format_pct(delta) if delta is not None else "N/A"
            elif m == "avg_sales_krw" or m.endswith("_krw"):
                fmt_start = _format_krw(start_val)
                fmt_end = _format_krw(end_val)
                fmt_delta = _format_krw(delta) if delta is not None else "N/A"
            else:
                fmt_start = _format_int(start_val)
                fmt_end = _format_int(end_val)
                fmt_delta = _format_int(delta) if delta is not None else "N/A"

            summary[m] = {
                "start_year": first["year"],
                "end_year": last["year"],
                "start_value": start_val,
                "end_value": end_val,
                "delta": delta,
                "trend": trend,
                "formatted": {
                    "start_value": fmt_start,
                    "end_value": fmt_end,
                    "delta": fmt_delta,
                },
            }

        return {
            "brand": {
                "brand_id": brand["brand_id"],
                "brand_name": brand["brand_name"],
                "company_name": brand["company_name"],
            },
            "range": {
                "start_year_requested": start_year,
                "end_year_requested": end_year,
                "start_year_used": timeline_rows[0]["year"],
                "end_year_used": timeline_rows[-1]["year"],
            },
            "metrics_used": sorted(metric_set),
            "timeline": timeline,
            "summary": summary,
        }
